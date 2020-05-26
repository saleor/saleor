import datetime
import shutil
from tempfile import NamedTemporaryFile
from unittest.mock import ANY, MagicMock, Mock, patch

import openpyxl
import petl as etl
import pytest
import pytz
from django.core.files import File
from freezegun import freeze_time

from saleor.core import JobStatus
from saleor.csv import ExportEvents, FileTypes
from saleor.csv.models import ExportEvent, ExportFile
from saleor.csv.utils.export import (
    append_to_file,
    create_csv_file_and_save_in_export_file,
    export_products,
    export_products_in_batches,
    get_filename,
    get_product_queryset,
    on_task_failure,
    on_task_success,
    save_csv_file_in_export_file,
)
from saleor.graphql.csv.enums import ProductFieldEnum
from saleor.product.models import Product


@patch("saleor.csv.utils.export.send_export_failed_info")
def test_on_task_failure(send_export_failed_info_mock, export_file):
    exc = Exception("Test")
    task_id = "task_id"
    args = [export_file.pk, {"all": ""}]
    kwargs = {}
    info_type = "Test error"
    info = Mock(type=info_type)

    assert export_file.status == JobStatus.PENDING
    assert export_file.created_at
    previous_updated_at = export_file.updated_at

    with freeze_time(datetime.datetime.now()) as frozen_datetime:
        on_task_failure(None, exc, task_id, args, kwargs, info)

        export_file.refresh_from_db()
        assert export_file.updated_at == pytz.utc.localize(frozen_datetime())

    assert export_file.updated_at != previous_updated_at
    assert export_file.status == JobStatus.FAILED
    assert export_file.created_at
    assert export_file.updated_at != previous_updated_at
    export_failed_event = ExportEvent.objects.get(
        export_file=export_file,
        user=export_file.created_by,
        type=ExportEvents.EXPORT_FAILED,
    )
    assert export_failed_event.parameters == {
        "message": str(exc),
        "error_type": info_type,
    }

    send_export_failed_info_mock.called_once_with(export_file, "export_failed")


def test_on_task_success(export_file):
    task_id = "task_id"
    args = [export_file.pk, {"filter": {}}]
    kwargs = {}

    assert export_file.status == JobStatus.PENDING
    assert export_file.created_at
    previous_updated_at = export_file.updated_at

    with freeze_time(datetime.datetime.now()) as frozen_datetime:
        on_task_success(None, None, task_id, args, kwargs)

        export_file.refresh_from_db()
        assert export_file.updated_at == pytz.utc.localize(frozen_datetime())
        assert export_file.updated_at != previous_updated_at

    assert export_file.status == JobStatus.SUCCESS
    assert export_file.created_at
    assert ExportEvent.objects.filter(
        export_file=export_file,
        user=export_file.created_by,
        type=ExportEvents.EXPORT_SUCCESS,
    )


@pytest.mark.parametrize(
    "scope, file_type",
    [
        ({"filter": {"is_published": True}}, FileTypes.CSV),
        ({"all": ""}, FileTypes.XLSX),
    ],
)
@patch("saleor.csv.utils.export.send_email_with_link_to_download_csv")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
@patch("saleor.csv.utils.export.export_products_in_batches")
def test_export_products(
    export_products_in_batches_mock,
    save_csv_file_in_export_file_mock,
    send_email_mock,
    product_list,
    export_file,
    scope,
    file_type,
):
    export_info = {
        "fields": [ProductFieldEnum.NAME.value],
        "warehouses": [],
        "attributes": [],
    }
    export_products(export_file.id, scope, export_info, file_type)

    export_products_in_batches_mock.called_once_with(
        ANY,
        export_info,
        {"id", "name"},
        ["id", "name"],
        {"id": "id", "name": "name"},
        ";",
        ANY,
        ANY,
        file_type,
    )
    save_csv_file_in_export_file_mock.called_once_with(export_file, ANY)
    send_email_mock.called_once_with(export_file)


@patch("saleor.csv.utils.export.send_email_with_link_to_download_csv")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_products_ids(
    save_csv_file_in_export_file_mock, send_email_mock, product_list, export_file
):
    pks = [product.pk for product in product_list[:2]]
    export_info = {"fields": [], "warehouses": [], "attributes": []}
    file_type = FileTypes.CSV

    assert export_file.status == JobStatus.PENDING
    assert not export_file.content_file

    export_products(export_file.id, {"ids": pks}, export_info, file_type)

    save_csv_file_in_export_file_mock.called_once_with(export_file, ANY)
    send_email_mock.called_once_with(export_file)


def test_get_filename_csv():
    with freeze_time("2000-02-09"):
        file_name = get_filename("test", FileTypes.CSV)

        assert file_name == "test_data_09_02_2000.csv"


def test_get_filename_xlsx():
    with freeze_time("2000-02-09"):
        file_name = get_filename("test", FileTypes.XLSX)

        assert file_name == "test_data_09_02_2000.xlsx"


def test_get_product_queryset_all(product_list):
    queryset = get_product_queryset({"all": ""})

    assert queryset.count() == len(product_list)


def test_get_product_queryset_ids(product_list):
    pks = [product.pk for product in product_list[:2]]
    queryset = get_product_queryset({"ids": pks})

    assert queryset.count() == len(pks)


def get_product_queryset_filter(product_list):
    product_not_published = product_list.first()
    product_not_published.is_published = False
    product_not_published.save()

    queryset = get_product_queryset({"ids": {"is_published": True}})

    assert queryset.count() == len(product_list) - 1


def test_create_csv_file_and_save_in_export_file_csv(export_file, tmpdir, media_root):
    export_data = [
        {"id": "123", "name": "test1", "collections": "coll1"},
        {"id": "345", "name": "test2"},
    ]
    headers = ["id", "name", "collections"]
    csv_headers_mapping = {"id": "ID", "name": "NAME"}
    delimiter = ";"
    export_file = export_file
    file_name = "test.csv"

    export_file_csv_upload_dir = ExportFile.content_file.field.upload_to

    assert not export_file.content_file

    create_csv_file_and_save_in_export_file(
        export_data,
        headers,
        csv_headers_mapping,
        delimiter,
        export_file,
        file_name,
        FileTypes.CSV,
    )

    csv_file = export_file.content_file
    assert csv_file
    assert csv_file.name == f"{export_file_csv_upload_dir}/{file_name}"

    file_content = csv_file.read().decode().split("\r\n")
    headers = list(csv_headers_mapping.values())
    headers.append("collections")

    assert ";".join(headers) in file_content
    assert ";".join(export_data[0].values()) in file_content
    assert (";".join(export_data[1].values()) + "; ") in file_content

    shutil.rmtree(tmpdir)


def test_create_csv_file_and_save_in_export_file_xlsx(export_file, tmpdir, media_root):
    export_data = [
        {"id": "123", "name": "test1", "collections": "coll1"},
        {"id": "345", "name": "test2"},
    ]
    headers = ["id", "name", "collections"]
    csv_headers_mapping = {"id": "ID", "name": "NAME", "collections": "COLLECTIONS"}
    delimiter = ";"
    file_name = "test.xlsx"

    export_file_csv_upload_dir = ExportFile.content_file.field.upload_to

    assert not export_file.content_file

    create_csv_file_and_save_in_export_file(
        export_data,
        headers,
        csv_headers_mapping,
        delimiter,
        export_file,
        file_name,
        FileTypes.XLSX,
    )

    xlsx_file = export_file.content_file
    assert xlsx_file
    assert xlsx_file.name == f"{export_file_csv_upload_dir}/{file_name}"

    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    expected_headers = list(csv_headers_mapping.values())
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]
    data = []
    for i in range(2, max_row + 1):
        row = []
        for j in range(1, max_col + 1):
            row.append(sheet_obj.cell(row=i, column=j).value)
        data.append(row)

    assert headers == expected_headers
    assert list(export_data[0].values()) in data
    row2 = list(export_data[1].values())
    # add string with space for collections column
    row2.append(" ")
    assert row2 in data

    shutil.rmtree(tmpdir)


def test_save_csv_file_in_export_file(export_file, tmpdir, media_root):
    file_mock = MagicMock(spec=File)
    file_mock.name = "temp_file.csv"
    file_name = "test.csv"

    assert not export_file.content_file

    save_csv_file_in_export_file(export_file, file_mock, file_name)

    export_file.refresh_from_db()
    assert export_file.content_file

    shutil.rmtree(tmpdir)


def test_append_to_file_for_csv(export_file, tmpdir, media_root):
    # given
    export_data = [
        {"id": "123", "name": "test1", "collections": "coll1"},
        {"id": "345", "name": "test2"},
    ]
    headers = ["id", "name", "collections"]
    delimiter = ";"

    file_name = "test.csv"

    table = etl.fromdicts([{"id": "1", "name": "A"}], header=headers, missing=" ")

    with NamedTemporaryFile() as temp_file:
        etl.tocsv(table, temp_file.name, delimiter=delimiter)
        export_file.content_file.save(file_name, temp_file)

    # when
    append_to_file(export_data, headers, export_file, FileTypes.CSV, delimiter)

    # then
    export_file.refresh_from_db()

    csv_file = export_file.content_file
    file_content = csv_file.read().decode().split("\r\n")
    assert ";".join(headers) in file_content
    assert ";".join(export_data[0].values()) in file_content
    assert (";".join(export_data[1].values()) + "; ") in file_content

    shutil.rmtree(tmpdir)


def test_append_to_file_for_xlsx(export_file, tmpdir, media_root):
    # given
    export_data = [
        {"id": "123", "name": "test1", "collections": "coll1"},
        {"id": "345", "name": "test2"},
    ]
    expected_headers = ["id", "name", "collections"]
    delimiter = ";"

    file_name = "test.xlsx"

    table = etl.fromdicts(
        [{"id": "1", "name": "A"}], header=expected_headers, missing=" "
    )

    with NamedTemporaryFile() as temp_file:
        etl.io.xlsx.toxlsx(table, temp_file.name)
        export_file.content_file.save(file_name, temp_file)

    # when
    append_to_file(
        export_data, expected_headers, export_file, FileTypes.XLSX, delimiter
    )

    # then
    export_file.refresh_from_db()

    xlsx_file = export_file.content_file
    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    expected_headers = expected_headers
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]
    data = []
    for i in range(2, max_row + 1):
        row = []
        for j in range(1, max_col + 1):
            row.append(sheet_obj.cell(row=i, column=j).value)
        data.append(row)

    assert headers == expected_headers
    assert list(export_data[0].values()) in data
    row2 = list(export_data[1].values())
    # add string with space for collections column
    row2.append(" ")
    assert row2 in data

    shutil.rmtree(tmpdir)


@patch("saleor.csv.utils.export.BATCH_SIZE", 1)
def test_export_products_in_batches_for_csv(
    product_list, export_file, tmpdir, media_root
):
    # given
    qs = Product.objects.all()
    export_info = {
        "fields": [ProductFieldEnum.NAME.value, ProductFieldEnum.VARIANT_SKU.value],
        "warehouses": [],
        "attributes": [],
    }
    export_fields = ["id", "name", "sku"]
    csv_headers_mapping = {"name": "name", "sku": "variant sku"}
    headers = ["id"] + list(csv_headers_mapping.keys())
    file_name = "test.csv"

    assert not export_file.content_file

    # when
    export_products_in_batches(
        qs,
        export_info,
        set(export_fields),
        headers,
        csv_headers_mapping,
        ";",
        export_file,
        file_name,
        FileTypes.CSV,
    )

    # then
    export_file.refresh_from_db()
    assert export_file.content_file

    expected_data = []
    for product in qs.order_by("pk"):
        product_data = []
        product_data.append(str(product.pk))
        product_data.append(product.name)

        for variant in product.variants.all():
            product_data.append(str(variant.sku))
            expected_data.append(product_data)

    csv_file = export_file.content_file
    file_content = csv_file.read().decode().split("\r\n")

    # ensure headers are in file
    assert ";".join(["id"] + list(csv_headers_mapping.values())) in file_content

    for row in expected_data:
        assert ";".join(row) in file_content

    shutil.rmtree(tmpdir)


@patch("saleor.csv.utils.export.BATCH_SIZE", 1)
def test_export_products_in_batches_for_xlsx(
    product_list, export_file, tmpdir, media_root
):
    # given
    qs = Product.objects.all()
    export_info = {
        "fields": [ProductFieldEnum.NAME.value, ProductFieldEnum.VARIANT_SKU.value],
        "warehouses": [],
        "attributes": [],
    }
    export_fields = ["id", "name", "sku"]
    csv_headers_mapping = {"name": "name", "sku": "variant sku"}
    headers = ["id"] + list(csv_headers_mapping.keys())
    file_name = "test.xlsx"

    assert not export_file.content_file

    # when
    export_products_in_batches(
        qs,
        export_info,
        set(export_fields),
        headers,
        csv_headers_mapping,
        ";",
        export_file,
        file_name,
        FileTypes.XLSX,
    )

    # then
    export_file.refresh_from_db()
    assert export_file.content_file

    expected_data = []
    for product in qs.order_by("pk"):
        product_data = []
        product_data.append(product.pk)
        product_data.append(product.name)

        for variant in product.variants.all():
            product_data.append(variant.sku)
            expected_data.append(product_data)

    xlsx_file = export_file.content_file
    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    expected_headers = ["id"] + list(csv_headers_mapping.values())
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]
    data = []
    for i in range(2, max_row + 1):
        row = []
        for j in range(1, max_col + 1):
            row.append(sheet_obj.cell(row=i, column=j).value)
        data.append(row)

    assert headers == expected_headers
    for row in expected_data:
        assert row in data

    shutil.rmtree(tmpdir)
