import datetime
import json
import shutil
from tempfile import NamedTemporaryFile
from unittest.mock import ANY, MagicMock, patch

import graphene
import openpyxl
import petl as etl
import pytest
from django.core.files import File
from freezegun import freeze_time

from ....core import JobStatus
from ....discount.models import VoucherCode
from ....giftcard.models import GiftCard
from ....graphql.csv.enums import ProductFieldEnum
from ....graphql.product.filters import ProductFilter
from ....product.models import Product, ProductChannelListing
from ... import FileTypes
from ...utils.export import (
    append_to_file,
    create_file_with_headers,
    export_gift_cards,
    export_gift_cards_in_batches,
    export_products,
    export_products_in_batches,
    export_voucher_codes,
    export_voucher_codes_in_batches,
    get_filename,
    get_queryset,
    parse_input,
    save_csv_file_in_export_file,
)


@pytest.mark.parametrize(
    "file_type",
    [FileTypes.CSV, FileTypes.XLSX],
)
@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_products_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_products(
    save_file_mock,
    send_email_mock,
    export_products_in_batches_mock,
    create_file_with_headers_mock,
    product_list,
    user_export_file,
    file_type,
):
    # given
    export_info = {
        "fields": [
            ProductFieldEnum.NAME.value,
            ProductFieldEnum.VARIANT_ID.value,
            ProductFieldEnum.VARIANT_SKU.value,
        ],
        "warehouses": [],
        "attributes": [],
        "channels": [],
    }

    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file

    product_list[0].variants.update(sku=None)

    # when
    export_products(user_export_file, {"all": ""}, export_info, file_type)

    # then
    create_file_with_headers_mock.assert_called_once_with(
        ["id", "name", "variant id", "variant sku"], ",", file_type
    )
    assert export_products_in_batches_mock.call_count == 1
    args, kwargs = export_products_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == set(
        Product.objects.all().values_list("pk", flat=True)
    )
    assert args[1:] == (
        export_info,
        {"id", "name", "variants__id", "variants__sku"},
        ["id", "name", "variants__id", "variants__sku"],
        ",",
        mock_file,
        file_type,
    )
    send_email_mock.assert_called_once_with(user_export_file, "products")
    save_file_mock.assert_called_once_with(user_export_file, mock_file, ANY)


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_products_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_products_ids(
    save_file_mock,
    send_email_mock,
    export_products_in_batches_mock,
    create_file_with_headers_mock,
    product_list,
    user_export_file,
):
    # given
    pks = [product.pk for product in product_list[:2]]
    export_info = {"fields": [], "warehouses": [], "attributes": [], "channels": []}
    file_type = FileTypes.CSV

    assert user_export_file.status == JobStatus.PENDING
    assert not user_export_file.content_file

    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file

    # when
    export_products(user_export_file, {"ids": pks}, export_info, file_type)

    # then
    create_file_with_headers_mock.assert_called_once_with(["id"], ",", file_type)

    assert export_products_in_batches_mock.call_count == 1
    args, kwargs = export_products_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == set(
        Product.objects.filter(pk__in=pks).values_list("pk", flat=True)
    )

    assert args[1:] == (
        export_info,
        {"id"},
        ["id"],
        ",",
        mock_file,
        file_type,
    )
    send_email_mock.assert_called_once_with(user_export_file, "products")
    save_file_mock.assert_called_once_with(user_export_file, mock_file, ANY)


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_products_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_products_filter_is_published(
    save_file_mock,
    send_email_mock,
    export_products_in_batches_mock,
    create_file_with_headers_mock,
    product_list,
    user_export_file,
    channel_USD,
):
    # given
    ProductChannelListing.objects.filter(
        product=product_list[0], channel=channel_USD
    ).update(is_published=False)

    export_info = {"fields": [], "warehouses": [], "attributes": []}
    file_type = FileTypes.CSV

    assert user_export_file.status == JobStatus.PENDING
    assert not user_export_file.content_file

    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file

    # when
    export_products(
        user_export_file,
        {"filter": {"is_published": True, "channel": channel_USD.slug}},
        export_info,
        file_type,
    )

    # then
    create_file_with_headers_mock.assert_called_once_with(["id"], ",", file_type)

    assert export_products_in_batches_mock.call_count == 1
    args, _ = export_products_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == set(
        Product.objects.filter(
            channel_listings__is_published=True, channel_listings__channel=channel_USD
        ).values_list("pk", flat=True)
    )
    assert args[1:] == (
        export_info,
        {"id"},
        ["id"],
        ",",
        mock_file,
        file_type,
    )
    send_email_mock.assert_called_once_with(user_export_file, "products")
    save_file_mock.assert_called_once_with(user_export_file, mock_file, ANY)


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_products_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_products_filter_collections(
    save_file_mock,
    send_email_mock,
    export_products_in_batches_mock,
    create_file_with_headers_mock,
    product_list,
    user_export_file,
    channel_USD,
    collection,
):
    # given
    collection.products.add(product_list[-1])

    export_info = {"fields": [], "warehouses": [], "attributes": []}
    file_type = FileTypes.CSV

    assert user_export_file.status == JobStatus.PENDING
    assert not user_export_file.content_file

    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file

    # when
    export_products(
        user_export_file,
        {
            "filter": {
                "collections": [graphene.Node.to_global_id("Collection", collection.pk)]
            }
        },
        export_info,
        file_type,
    )

    # then
    create_file_with_headers_mock.assert_called_once_with(["id"], ",", file_type)

    assert export_products_in_batches_mock.call_count == 1
    batch_args, _ = export_products_in_batches_mock.call_args
    assert set(batch_args[0].values_list("pk", flat=True)) == {product_list[-1].pk}
    assert batch_args[1:] == (export_info, {"id"}, ["id"], ",", mock_file, file_type)
    send_email_mock.assert_called_once_with(user_export_file, "products")
    save_file_mock.assert_called_once_with(user_export_file, mock_file, ANY)


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_products_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_products_by_app(
    save_file_mock,
    send_email_mock,
    export_products_in_batches_mock,
    create_file_with_headers_mock,
    product_list,
    app_export_file,
):
    # given
    export_info = {
        "fields": [ProductFieldEnum.NAME.value],
        "warehouses": [],
        "attributes": [],
        "channels": [],
    }
    file_type = FileTypes.CSV

    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file

    # when
    export_products(app_export_file, {"all": ""}, export_info, file_type)

    # then
    create_file_with_headers_mock.assert_called_once_with(
        ["id", "name"], ",", file_type
    )

    assert export_products_in_batches_mock.call_count == 1
    args, kwargs = export_products_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == set(
        Product.objects.all().values_list("pk", flat=True)
    )
    assert args[1:] == (
        export_info,
        {"id", "name"},
        ["id", "name"],
        ",",
        mock_file,
        file_type,
    )

    send_email_mock.assert_called_once_with(app_export_file, "products")

    save_file_mock.assert_called_once_with(app_export_file, mock_file, ANY)


@patch("saleor.plugins.manager.PluginsManager.product_export_completed")
def test_export_products_webhook(
    mocked_product_export_completed,
    product_list,
    user_export_file,
    media_root,
):
    # given
    product_list[0].variants.update(sku=None)

    # when
    export_products(user_export_file, {"all": ""}, {}, FileTypes.CSV)

    # then
    mocked_product_export_completed.assert_called_once_with(user_export_file)


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_gift_cards_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_gift_cards(
    save_file_mock,
    send_email_mock,
    export_in_batches_mock,
    create_file_with_headers_mock,
    user_export_file,
    gift_card,
    gift_card_expiry_date,
    gift_card_used,
):
    # given
    file_type = FileTypes.CSV

    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file

    # when
    export_gift_cards(user_export_file, {"all": ""}, file_type)

    # then
    create_file_with_headers_mock.assert_called_once_with(["code"], ",", file_type)

    assert export_in_batches_mock.call_count == 1
    args, kwargs = export_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == set(
        GiftCard.objects.exclude(id=gift_card_used.id).values_list("pk", flat=True)
    )
    assert args[1:] == (
        ["code"],
        ",",
        mock_file,
        file_type,
    )

    send_email_mock.assert_called_once_with(user_export_file, "gift cards")

    save_file_mock.assert_called_once_with(user_export_file, mock_file, ANY)


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_gift_cards_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_gift_cards_by_app(
    save_file_mock,
    send_email_mock,
    export_in_batches_mock,
    create_file_with_headers_mock,
    app_export_file,
    gift_card,
    gift_card_expiry_date,
    gift_card_used,
):
    file_type = FileTypes.CSV

    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file

    # when
    export_gift_cards(app_export_file, {"all": ""}, file_type)

    # then
    create_file_with_headers_mock.assert_called_once_with(["code"], ",", file_type)

    assert export_in_batches_mock.call_count == 1
    args, kwargs = export_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == set(
        GiftCard.objects.exclude(id=gift_card_used.id).values_list("pk", flat=True)
    )
    assert args[1:] == (
        ["code"],
        ",",
        mock_file,
        file_type,
    )

    send_email_mock.assert_called_once_with(app_export_file, "gift cards")

    save_file_mock.assert_called_once_with(app_export_file, mock_file, ANY)


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_gift_cards_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_gift_cards_ids(
    save_file_mock,
    send_email_mock,
    export_in_batches_mock,
    create_file_with_headers_mock,
    user_export_file,
    gift_card,
    gift_card_expiry_date,
    gift_card_used,
):
    file_type = FileTypes.CSV

    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file
    pks = [gift_card.pk]

    # when
    export_gift_cards(user_export_file, {"ids": pks}, file_type)

    # then
    create_file_with_headers_mock.assert_called_once_with(["code"], ",", file_type)

    assert export_in_batches_mock.call_count == 1
    args, kwargs = export_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == set(pks)
    assert args[1:] == (
        ["code"],
        ",",
        mock_file,
        file_type,
    )

    send_email_mock.assert_called_once_with(user_export_file, "gift cards")

    save_file_mock.assert_called_once_with(user_export_file, mock_file, ANY)


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_gift_cards_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_gift_cards_with_filter(
    save_file_mock,
    send_email_mock,
    export_in_batches_mock,
    create_file_with_headers_mock,
    user_export_file,
    gift_card,
    gift_card_expiry_date,
    gift_card_used,
    shippable_gift_card_product,
):
    file_type = FileTypes.CSV

    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file

    gift_card_expiry_date.product = shippable_gift_card_product
    gift_card_used.product = shippable_gift_card_product
    GiftCard.objects.bulk_update([gift_card_expiry_date, gift_card_used], ["product"])

    # when
    export_gift_cards(
        user_export_file,
        {
            "filter": {
                "products": [
                    graphene.Node.to_global_id(
                        "Product", shippable_gift_card_product.pk
                    )
                ]
            }
        },
        file_type,
    )

    # then
    create_file_with_headers_mock.assert_called_once_with(["code"], ",", file_type)

    assert export_in_batches_mock.call_count == 1
    args, kwargs = export_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == {gift_card_expiry_date.pk}
    assert args[1:] == (
        ["code"],
        ",",
        mock_file,
        file_type,
    )

    send_email_mock.assert_called_once_with(user_export_file, "gift cards")

    save_file_mock.assert_called_once_with(user_export_file, mock_file, ANY)


@patch("saleor.plugins.manager.PluginsManager.gift_card_export_completed")
def test_export_gift_cards_webhook(
    mocked_gift_card_export_completed,
    user_export_file,
    gift_card,
    gift_card_expiry_date,
    gift_card_used,
    media_root,
):
    # given
    file_type = FileTypes.CSV

    # when
    export_gift_cards(user_export_file, {"all": ""}, file_type)

    # then
    mocked_gift_card_export_completed.assert_called_once_with(user_export_file)


def test_get_filename_csv():
    with freeze_time("2000-02-09 03:21:34"):
        file_name = get_filename("test", FileTypes.CSV)

        assert file_name.startswith("test_data_09_02_2000_03_21_34")
        assert file_name.endswith(".csv")


def test_get_filename_xlsx():
    with freeze_time("2000-02-09 05:22:44"):
        file_name = get_filename("test", FileTypes.XLSX)

        assert file_name.startswith("test_data_09_02_2000_05_22_44")
        assert file_name.endswith(".xlsx")


def test_get_product_queryset_all(product_list):
    queryset = get_queryset(Product, ProductFilter, {"all": ""})

    assert queryset.count() == len(product_list)


def test_get_product_queryset_ids(product_list):
    pks = [product.pk for product in product_list[:2]]
    queryset = get_queryset(Product, ProductFilter, {"ids": pks})

    assert queryset.count() == len(pks)


def get_product_queryset_filter(product_list):
    product_not_published = product_list.first()
    product_not_published.is_published = False
    product_not_published.save()

    queryset = get_queryset(Product, ProductFilter, {"ids": {"is_published": True}})

    assert queryset.count() == len(product_list) - 1


def test_create_file_with_headers_csv(user_export_file, tmpdir, media_root):
    # given
    file_headers = ["id", "name", "collections"]

    assert not user_export_file.content_file

    # when
    csv_file = create_file_with_headers(file_headers, ",", FileTypes.CSV)

    # then
    assert csv_file

    file_content = csv_file.read().decode().split("\r\n")

    assert ",".join(file_headers) in file_content

    shutil.rmtree(tmpdir)


def test_create_file_with_headers_xlsx(user_export_file, tmpdir, media_root):
    # given
    file_headers = ["id", "name", "collections"]

    assert not user_export_file.content_file

    # when
    xlsx_file = create_file_with_headers(file_headers, ",", FileTypes.XLSX)

    # then
    assert xlsx_file

    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]

    assert headers == file_headers

    shutil.rmtree(tmpdir)


def test_save_csv_file_in_export_file(user_export_file, tmpdir, media_root):
    file_mock = MagicMock(spec=File)
    file_mock.name = "temp_file.csv"
    file_name = "test.csv"

    assert not user_export_file.content_file

    save_csv_file_in_export_file(user_export_file, file_mock, file_name)

    user_export_file.refresh_from_db()
    assert user_export_file.content_file

    shutil.rmtree(tmpdir)


def test_append_to_file_for_csv(user_export_file, tmpdir, media_root):
    # given
    export_data = [
        {"id": "123", "name": "test1", "collections": "coll1"},
        {"id": "345", "name": "test2"},
    ]
    headers = ["id", "name", "collections"]
    delimiter = ","

    table = etl.fromdicts([{"id": "1", "name": "A"}], header=headers, missing="")

    temp_file = NamedTemporaryFile()
    etl.tocsv(table, temp_file.name, delimiter=delimiter)

    # when
    append_to_file(export_data, headers, temp_file, FileTypes.CSV, delimiter)

    # then
    user_export_file.refresh_from_db()

    file_content = temp_file.read().decode().split("\r\n")
    assert ",".join(headers) in file_content
    assert ",".join(export_data[0].values()) in file_content
    assert (",".join(export_data[1].values()) + ",") in file_content

    temp_file.close()
    shutil.rmtree(tmpdir)


def test_append_to_file_for_xlsx(user_export_file, tmpdir, media_root):
    # given
    export_data = [
        {"id": "123", "name": "test1", "collections": "coll1"},
        {"id": "345", "name": "test2"},
    ]
    expected_headers = ["id", "name", "collections"]

    table = etl.fromdicts(
        [{"id": "1", "name": "A"}], header=expected_headers, missing=""
    )

    temp_file = NamedTemporaryFile(suffix=".xlsx")
    etl.io.xlsx.toxlsx(table, temp_file.name)

    # when
    append_to_file(export_data, expected_headers, temp_file, FileTypes.XLSX, ",")

    # then
    user_export_file.refresh_from_db()

    workbook = openpyxl.load_workbook(temp_file)

    sheet = workbook.worksheets[0]
    assert sheet.cell(1, 1).value == expected_headers[0]
    assert sheet.cell(1, 2).value == expected_headers[1]
    assert sheet.cell(1, 3).value == expected_headers[2]
    assert sheet.cell(3, 1).value == export_data[0]["id"]
    assert sheet.cell(3, 2).value == export_data[0]["name"]
    assert sheet.cell(3, 3).value == export_data[0]["collections"]
    assert sheet.cell(4, 1).value == export_data[1]["id"]
    assert sheet.cell(4, 2).value == export_data[1]["name"]
    assert sheet.cell(4, 3).value is None

    temp_file.close()
    shutil.rmtree(tmpdir)


@patch("saleor.csv.utils.export.BATCH_SIZE", 1)
def test_export_products_in_batches_for_csv(
    product_list,
    user_export_file,
    tmpdir,
    media_root,
):
    # given
    qs = Product.objects.all().order_by("pk")
    export_info = {
        "fields": [
            ProductFieldEnum.NAME.value,
            ProductFieldEnum.DESCRIPTION.value,
            ProductFieldEnum.VARIANT_SKU.value,
        ],
        "warehouses": [],
        "attributes": [],
        "channels": [],
    }
    export_fields = ["id", "name", "variants__sku"]
    expected_headers = ["id", "name", "variant sku"]

    table = etl.wrap([expected_headers])

    temp_file = NamedTemporaryFile()
    etl.tocsv(table, temp_file.name, delimiter=",")

    # when
    export_products_in_batches(
        qs,
        export_info,
        set(export_fields),
        export_fields,
        ",",
        temp_file,
        FileTypes.CSV,
    )

    # then

    expected_data = []
    for product in qs.order_by("pk"):
        product_data = []
        id = graphene.Node.to_global_id("Product", product.pk)
        product_data.append(id)
        product_data.append(product.name)

        for variant in product.variants.all():
            product_data.append(str(variant.sku))
            expected_data.append(product_data)

    file_content = temp_file.read().decode().split("\r\n")

    # ensure headers are in file
    assert ",".join(expected_headers) in file_content

    for row in expected_data:
        assert ",".join(row) in file_content

    shutil.rmtree(tmpdir)


@patch("saleor.csv.utils.export.BATCH_SIZE", 1)
def test_export_products_in_batches_for_xlsx(
    product_list,
    user_export_file,
    tmpdir,
    media_root,
):
    # given
    product = product_list[0]
    product.description = {
        "blocks": [
            {"data": {"text": "This is an example description."}, "type": "paragraph"}
        ]
    }
    product.save(update_fields=["description"])

    qs = Product.objects.all().order_by("pk")
    export_info = {
        "fields": [
            ProductFieldEnum.NAME.value,
            ProductFieldEnum.DESCRIPTION.value,
            ProductFieldEnum.VARIANT_SKU.value,
        ],
        "warehouses": [],
        "attributes": [],
        "channels": [],
    }
    export_fields = ["id", "name", "description_as_str", "variants__sku"]
    expected_headers = ["id", "name", "description", "variant sku"]

    table = etl.wrap([expected_headers])

    temp_file = NamedTemporaryFile(suffix=".xlsx")
    etl.io.xlsx.toxlsx(table, temp_file.name)

    # when
    export_products_in_batches(
        qs,
        export_info,
        set(export_fields),
        export_fields,
        ",",
        temp_file,
        FileTypes.XLSX,
    )

    # then
    expected_data = []
    for product in qs:
        product_data = []
        id = graphene.Node.to_global_id("Product", product.pk)
        product_data.append(id)
        product_data.append(product.name)
        product_data.append(json.dumps(product.description))

        for variant in product.variants.all():
            product_data.append(variant.sku)
            expected_data.append(product_data)

    wb_obj = openpyxl.load_workbook(temp_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]
    data = []
    for i in range(2, max_row + 1):
        row = []
        for j in range(1, max_col + 1):
            row.append(sheet_obj.cell(row=i, column=j).value)
        data.append(row)

    assert headers == expected_headers
    for row in expected_data:
        assert row in data

    shutil.rmtree(tmpdir)


@patch("saleor.csv.utils.export.BATCH_SIZE", 1)
def test_export_gift_cards_in_batches_to_csv(
    gift_card,
    gift_card_expiry_date,
    gift_card_used,
    tmpdir,
):
    # given
    gift_cards = GiftCard.objects.exclude(id=gift_card_used.id).order_by("pk")

    table = etl.wrap([["code"]])
    temp_file = NamedTemporaryFile()
    etl.tocsv(table, temp_file.name, delimiter=",")

    # when
    export_gift_cards_in_batches(
        gift_cards,
        ["code"],
        ",",
        temp_file,
        "csv",
    )

    # then
    file_content = temp_file.read().decode().split("\r\n")

    # ensure headers are in the file
    assert "code" in file_content

    for card in gift_cards:
        assert card.code in file_content

    shutil.rmtree(tmpdir)


@patch("saleor.csv.utils.export.BATCH_SIZE", 1)
def test_export_gift_cards_in_batches_to_xlsx(
    gift_card,
    gift_card_expiry_date,
    gift_card_used,
    tmpdir,
):
    # given
    gift_cards = GiftCard.objects.exclude(id=gift_card_used.id).order_by("pk")

    table = etl.wrap([["code"]])
    temp_file = NamedTemporaryFile(suffix=".xlsx")
    etl.io.xlsx.toxlsx(table, temp_file.name)

    # when
    export_gift_cards_in_batches(
        gift_cards,
        ["code"],
        ",",
        temp_file,
        "xlsx",
    )

    # then
    wb_obj = openpyxl.load_workbook(temp_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]
    data = []
    for i in range(2, max_row + 1):
        row = []
        for j in range(1, max_col + 1):
            row.append(sheet_obj.cell(row=i, column=j).value)
        data.append(row)

    assert headers == ["code"]
    for card in gift_cards:
        assert [card.code] in data

    shutil.rmtree(tmpdir)


def test_parse_input():
    data = {
        "collections": None,
        "categories": None,
        "attributes": [
            {
                "slug": "release-date-time",
                "date_time": {
                    "gte": "2019-08-08T00:00:00+02:00",
                    "lte": "2021-08-08T00:00:00+02:00",
                },
            },
            {
                "slug": "release-date",
                "date": {
                    "gte": "2019-08-08",
                },
            },
        ],
        "stock_availability": None,
        "price": None,
        "product_types": None,
    }
    parsed_data = parse_input(data)

    assert isinstance(
        parsed_data["attributes"][0]["date_time"]["gte"], datetime.datetime
    )
    assert isinstance(
        parsed_data["attributes"][0]["date_time"]["lte"], datetime.datetime
    )
    assert isinstance(parsed_data["attributes"][1]["date"]["gte"], datetime.date)

    data.pop("attributes")
    parsed_data = parse_input(data)

    assert data == parsed_data


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_voucher_codes_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_voucher_codes_by_voucher_id(
    save_file_mock,
    send_email_mock,
    export_in_batches_mock,
    create_file_with_headers_mock,
    user_export_file,
    voucher_with_many_codes,
    voucher_percentage,
):
    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file
    file_type = FileTypes.CSV
    voucher = voucher_with_many_codes

    # when
    export_voucher_codes(user_export_file, file_type, voucher_id=voucher.id)

    # then
    create_file_with_headers_mock.assert_called_once_with(["code"], ",", file_type)

    assert export_in_batches_mock.call_count == 1
    args, kwargs = export_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == set(
        VoucherCode.objects.filter(voucher_id=voucher.id).values_list("pk", flat=True)
    )
    assert args[1:] == (
        ["code"],
        ",",
        mock_file,
        file_type,
    )

    send_email_mock.assert_called_once_with(user_export_file, "voucher codes")

    save_file_mock.assert_called_once_with(user_export_file, mock_file, ANY)


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_voucher_codes_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_voucher_codes_by_ids(
    save_file_mock,
    send_email_mock,
    export_in_batches_mock,
    create_file_with_headers_mock,
    user_export_file,
    voucher_with_many_codes,
    voucher_percentage,
):
    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file
    file_type = FileTypes.CSV
    voucher = voucher_with_many_codes
    code_ids = [code.id for code in voucher.codes.all()]

    # when
    export_voucher_codes(user_export_file, file_type, ids=code_ids)

    # then
    create_file_with_headers_mock.assert_called_once_with(["code"], ",", file_type)

    assert export_in_batches_mock.call_count == 1
    args, kwargs = export_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == set(
        VoucherCode.objects.filter(id__in=code_ids).values_list("pk", flat=True)
    )
    assert args[1:] == (
        ["code"],
        ",",
        mock_file,
        file_type,
    )

    send_email_mock.assert_called_once_with(user_export_file, "voucher codes")

    save_file_mock.assert_called_once_with(user_export_file, mock_file, ANY)


@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_voucher_codes_in_batches")
@patch("saleor.csv.utils.export.send_export_download_link_notification")
@patch("saleor.csv.utils.export.save_csv_file_in_export_file")
def test_export_voucher_codes_by_app(
    save_file_mock,
    send_email_mock,
    export_in_batches_mock,
    create_file_with_headers_mock,
    app_export_file,
    voucher_with_many_codes,
):
    mock_file = MagicMock(spec=File)
    create_file_with_headers_mock.return_value = mock_file
    file_type = FileTypes.CSV
    voucher = voucher_with_many_codes

    # when
    export_voucher_codes(app_export_file, file_type, voucher_id=voucher.id)

    # then
    create_file_with_headers_mock.assert_called_once_with(["code"], ",", file_type)

    assert export_in_batches_mock.call_count == 1
    args, kwargs = export_in_batches_mock.call_args
    assert set(args[0].values_list("pk", flat=True)) == set(
        VoucherCode.objects.filter(voucher_id=voucher.id).values_list("pk", flat=True)
    )
    assert args[1:] == (
        ["code"],
        ",",
        mock_file,
        file_type,
    )

    send_email_mock.assert_called_once_with(app_export_file, "voucher codes")

    save_file_mock.assert_called_once_with(app_export_file, mock_file, ANY)


@patch("saleor.plugins.manager.PluginsManager.voucher_code_export_completed")
def test_export_voucher_codes_webhooks(
    mocked_voucher_code_export_completed,
    user_export_file,
    voucher_with_many_codes,
    media_root,
):
    file_type = FileTypes.CSV
    voucher = voucher_with_many_codes

    # when
    export_voucher_codes(user_export_file, file_type, voucher_id=voucher.id)

    # then
    mocked_voucher_code_export_completed.assert_called_once()


@patch("saleor.csv.utils.export.BATCH_SIZE", 1)
def test_export_voucher_codes_in_batches_to_csv(
    voucher_with_many_codes,
    tmpdir,
):
    # given
    voucher_codes = voucher_with_many_codes.codes.all()

    table = etl.wrap([["code"]])
    temp_file = NamedTemporaryFile()
    etl.tocsv(table, temp_file.name, delimiter=",")

    # when
    export_voucher_codes_in_batches(
        voucher_codes,
        ["code"],
        ",",
        temp_file,
        "csv",
    )

    # then
    file_content = temp_file.read().decode().split("\r\n")

    # ensure headers are in the file
    assert "code" in file_content

    for voucher_code in voucher_codes:
        assert voucher_code.code in file_content

    shutil.rmtree(tmpdir)


@patch("saleor.csv.utils.export.BATCH_SIZE", 1)
def test_export_voucher_codes_in_batches_to_xlsx(
    voucher_with_many_codes,
    tmpdir,
):
    # given
    voucher_codes = voucher_with_many_codes.codes.all()

    table = etl.wrap([["code"]])
    temp_file = NamedTemporaryFile(suffix=".xlsx")
    etl.io.xlsx.toxlsx(table, temp_file.name)

    # when
    export_voucher_codes_in_batches(
        voucher_codes,
        ["code"],
        ",",
        temp_file,
        "xlsx",
    )

    # then
    wb_obj = openpyxl.load_workbook(temp_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]
    data = []
    for i in range(2, max_row + 1):
        row = []
        for j in range(1, max_col + 1):
            row.append(sheet_obj.cell(row=i, column=j).value)
        data.append(row)

    assert headers == ["code"]
    for voucher_code in voucher_codes:
        assert [voucher_code.code] in data

    shutil.rmtree(tmpdir)
