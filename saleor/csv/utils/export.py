import datetime
import logging
import uuid
from decimal import Decimal
from tempfile import NamedTemporaryFile
from typing import IO, TYPE_CHECKING, Any, cast

import openpyxl
import petl as etl
from django.conf import settings
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ...core.db.connection import allow_writer
from ...core.utils.batches import queryset_in_batches
from ...discount.models import VoucherCode
from ...giftcard.models import GiftCard
from ...inventory.models import PurchaseOrder
from ...order.models import Order
from ...product.models import Product
from .. import FileTypes
from ..notifications import send_export_download_link_notification
from .image_embedding import embed_images_in_excel
from .product_headers import get_product_export_fields_and_headers_info
from .products_data import get_products_data
from .variant_compression import get_products_data_compressed

logger = logging.getLogger(__name__)


def convert_ids_to_proper_type(ids: list[str]) -> list[uuid.UUID | int]:
    """Convert string IDs to UUID or int based on format.

    Args:
        ids: List of string IDs

    Returns:
        List of converted IDs (UUID or int)

    """
    converted: list[uuid.UUID | int] = []
    for id_value in ids:
        if isinstance(id_value, str):
            # Try to convert to UUID, fall back to int
            try:
                converted.append(uuid.UUID(id_value))
            except ValueError:
                converted.append(int(id_value))
        else:
            converted.append(id_value)
    return converted


if TYPE_CHECKING:
    from django.db.models import QuerySet

    from ..models import ExportFile


BATCH_SIZE = 1000


def export_products(
    export_file: "ExportFile",
    scope: dict[str, str | dict],
    export_info: dict[str, list],
    file_type: str,
    delimiter: str = ",",
):
    from ...graphql.product.filters.product import ProductFilter

    file_name = get_filename("product", file_type)
    queryset = get_queryset(Product, ProductFilter, scope)

    (
        export_fields,
        file_headers,
        data_headers,
    ) = get_product_export_fields_and_headers_info(export_info)

    temporary_file = create_file_with_headers(file_headers, delimiter, file_type)

    # Check if we should compress variants
    compress_variants = export_info.get("compress_variants", False)

    if compress_variants:
        # Export with compressed variants (one row per product)
        export_products_compressed(
            queryset,
            export_info,
            data_headers,
            delimiter,
            temporary_file,
            file_type,
        )
    else:
        # Export with expanded variants (one row per variant)
        export_products_in_batches(
            queryset,
            export_info,
            set(export_fields),
            data_headers,
            delimiter,
            temporary_file,
            file_type,
        )

    # Embed images in Excel file if requested
    embed_images = export_info.get("embed_images", False)
    logger.info(
        "Image embedding: embed_images=%s, file_type=%s, file_headers=%s",
        embed_images,
        file_type,
        file_headers,
    )

    if embed_images and file_type == FileTypes.XLSX:
        # Identify which columns contain images
        image_columns = []
        if "product media" in file_headers:
            image_columns.append("product media")
        # Only include variant media if NOT compressing variants
        # (variant media doesn't make sense when there's one row per product)
        if "variant media" in file_headers and not compress_variants:
            image_columns.append("variant media")

        logger.info("Image columns to embed: %s", image_columns)

        # Embed images if there are any image columns
        if image_columns:
            logger.info("Starting image embedding for %s", temporary_file.name)
            embed_images_in_excel(temporary_file.name, image_columns)
            logger.info("Image embedding completed")
        else:
            logger.warning("No image columns found to embed")

    # Format currency columns in Excel file
    if file_type == FileTypes.XLSX and export_info.get("channels"):
        logger.info("Applying currency formatting to price columns")
        format_currency_columns(temporary_file.name, export_info)

    # Apply price list formatting if requested
    if file_type == FileTypes.XLSX and export_info.get("price_list_format"):
        logger.info("Applying price list formatting")
        format_as_price_list(temporary_file.name, export_info)

    save_csv_file_in_export_file(export_file, temporary_file, file_name)
    temporary_file.close()
    send_export_download_link_notification(export_file, "products")


def export_gift_cards(
    export_file: "ExportFile",
    scope: dict[str, str | dict],
    file_type: str,
    delimiter: str = ",",
):
    from ...graphql.giftcard.filters import GiftCardFilter

    file_name = get_filename("gift_card", file_type)

    queryset = get_queryset(GiftCard, GiftCardFilter, scope)
    # only unused gift cards codes can be exported
    queryset = queryset.filter(used_by_email__isnull=True)

    export_fields = ["code"]
    temporary_file = create_file_with_headers(export_fields, delimiter, file_type)

    export_gift_cards_in_batches(
        queryset,
        export_fields,
        delimiter,
        temporary_file,
        file_type,
    )

    save_csv_file_in_export_file(export_file, temporary_file, file_name)
    temporary_file.close()
    send_export_download_link_notification(export_file, "gift cards")


ORDER_SUMMARY_HEADERS = [
    "Number",
    "Date",
    "Status",
    "Charge Status",
    "Customer Name",
    "Customer Email",
    "Billing Address",
    "Shipping Address",
    "Channel",
    "Currency",
    "Shipping Method",
    "Shipping Net",
    "Shipping Gross",
    "Incoterm",
    "Voucher Code",
    "Total Net",
    "Total Gross",
]

ORDER_LINE_HEADERS = [
    "Number",
    "Product Name",
    "Variant Name",
    "SKU",
    "Product Code",
    "Quantity",
    "Unit Price Net",
    "Unit Price Gross",
    "Line Total Net",
    "Line Total Gross",
    "Tax Rate",
    "Tax Class",
    "Tax Class Country",
    "Allocation Sources",
]

# Flat/denormalized headers for CSV (order fields + line fields, no duplicate Number)
ORDER_EXPORT_HEADERS = ORDER_SUMMARY_HEADERS + [
    h for h in ORDER_LINE_HEADERS if h != "Number"
]


def _format_address(address) -> str:
    if not address:
        return ""
    parts = [
        f"{address.first_name} {address.last_name}".strip(),
        address.company_name,
        address.street_address_1,
        address.street_address_2,
        address.city,
        address.city_area,
        address.postal_code,
        address.country_area,
        str(address.country) if address.country else "",
    ]
    return ", ".join(p for p in parts if p)


def _build_order_rows(order, product_code_slug: str) -> tuple[dict, list[dict]]:
    billing = _format_address(order.billing_address)
    shipping_addr = _format_address(order.shipping_address)
    customer_name = ""
    if order.billing_address:
        customer_name = f"{order.billing_address.first_name} {order.billing_address.last_name}".strip()

    summary: dict = {
        "Number": str(order.number),
        "Date": order.created_at.strftime("%Y-%m-%d %H:%M:%S")
        if order.created_at
        else "",
        "Status": order.status,
        "Charge Status": order.charge_status,
        "Customer Name": customer_name,
        "Customer Email": order.user_email or "",
        "Billing Address": billing,
        "Shipping Address": shipping_addr,
        "Channel": order.channel.name if order.channel_id else "",
        "Currency": order.currency,
        "Shipping Method": order.shipping_method_name or "",
        "Shipping Net": str(order.shipping_price_net_amount),
        "Shipping Gross": str(order.shipping_price_gross_amount),
        "Incoterm": order.inco_term or "",
        "Voucher Code": order.voucher_code or "",
        "Total Net": str(order.total_net_amount),
        "Total Gross": str(order.total_gross_amount),
    }

    line_rows: list[dict] = []
    for line in order.lines.all():
        tax_class_country = ""
        if line.tax_class_country_rate_id and line.tax_class_country_rate:
            tax_class_country = str(line.tax_class_country_rate.country)
        product_code = ""
        if line.variant_id and line.variant:
            for av in line.variant.product.attributevalues.all():
                if av.value.attribute.slug == product_code_slug:
                    product_code = av.value.name
                    break

        allocation_details = []
        for allocation in line.allocations.all():
            for source in allocation.allocation_sources.all():
                poi = source.purchase_order_item
                po = poi.order
                po_name = po.name or str(po.pk)
                supplier = po.source_warehouse.name
                unit_price = _compute_unit_price(poi)
                price = str(round(unit_price, 2)) if unit_price else "?"
                currency = poi.currency or ""
                coo = str(poi.country_of_origin) if poi.country_of_origin else ""
                parts = [f"{source.quantity}@{price} {currency}".strip(), supplier]
                if coo:
                    parts.append(coo)
                allocation_details.append(f"{po_name}({', '.join(parts)})")

        line_rows.append(
            {
                "Number": str(order.number),
                "Product Name": line.product_name,
                "Variant Name": line.variant_name,
                "SKU": line.product_sku or "",
                "Product Code": product_code,
                "Quantity": line.quantity,
                "Unit Price Net": str(line.unit_price_net_amount),
                "Unit Price Gross": str(line.unit_price_gross_amount),
                "Line Total Net": str(line.total_price_net_amount),
                "Line Total Gross": str(line.total_price_gross_amount),
                "Tax Rate": str(line.tax_rate) if line.tax_rate is not None else "",
                "Tax Class": line.tax_class_name or "",
                "Tax Class Country": tax_class_country,
                "Allocation Sources": "; ".join(allocation_details),
            }
        )

    return summary, line_rows


def _order_batch_queryset(batch_pks):
    return (
        Order.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME)
        .filter(pk__in=batch_pks)
        .select_related(
            "billing_address",
            "shipping_address",
            "channel",
            "shipping_method",
            "user",
        )
        .prefetch_related(
            "lines",
            "lines__tax_class_country_rate",
            "lines__variant__product__attributevalues__value__attribute",
            "lines__allocations__allocation_sources__purchase_order_item__order__source_warehouse",
            "lines__allocations__allocation_sources__purchase_order_item__adjustments",
        )
        .order_by("pk")
    )


def export_orders(
    export_file: "ExportFile",
    scope: dict[str, str | dict],
    file_type: str,
    delimiter: str = ",",
):
    from ...graphql.order.filters import OrderFilter
    from ...site.models import SiteSettings

    file_name = get_filename("order", file_type)
    queryset = get_queryset(Order, OrderFilter, scope)

    site_settings = SiteSettings.objects.first()
    product_code_slug = (
        site_settings.invoice_product_code_attribute
        if site_settings
        else "product-code"
    )

    if file_type == FileTypes.XLSX:
        summary_rows: list[dict] = []
        lines_rows: list[dict] = []
        for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
            for order in _order_batch_queryset(batch_pks):
                summary, line_rows = _build_order_rows(order, product_code_slug)
                summary_rows.append(summary)
                lines_rows.extend(line_rows)

        temp_file = NamedTemporaryFile("ab+", suffix=".xlsx")
        _write_orders_xlsx(summary_rows, lines_rows, temp_file.name)
        temporary_file = temp_file
    else:
        temporary_file = create_file_with_headers(
            ORDER_EXPORT_HEADERS, delimiter, file_type
        )
        export_orders_in_batches(queryset, delimiter, temporary_file, product_code_slug)

    save_csv_file_in_export_file(export_file, temporary_file, file_name)
    temporary_file.close()
    send_export_download_link_notification(export_file, "orders")


def _write_orders_xlsx(summary_rows: list[dict], lines_rows: list[dict], filename: str):
    wb = openpyxl.Workbook()

    ws_orders = wb.active
    ws_orders.title = "Orders"
    ws_orders.append(ORDER_SUMMARY_HEADERS)
    for row in summary_rows:
        ws_orders.append([row.get(h, "") for h in ORDER_SUMMARY_HEADERS])

    ws_lines = wb.create_sheet("Lines")
    ws_lines.append(ORDER_LINE_HEADERS)
    for row in lines_rows:
        ws_lines.append([row.get(h, "") for h in ORDER_LINE_HEADERS])

    wb.save(filename)


def export_orders_in_batches(
    queryset: "QuerySet",
    delimiter: str,
    temporary_file: Any,
    product_code_slug: str = "product-code",
):
    for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
        export_data: list[dict] = []
        for order in _order_batch_queryset(batch_pks):
            summary, line_rows = _build_order_rows(order, product_code_slug)
            if not line_rows:
                export_data.append(
                    {**summary, **{h: "" for h in ORDER_LINE_HEADERS if h != "Number"}}
                )
            else:
                for line_row in line_rows:
                    export_data.append(
                        {
                            **summary,
                            **{k: v for k, v in line_row.items() if k != "Number"},
                        }
                    )
        append_to_file(
            export_data, ORDER_EXPORT_HEADERS, temporary_file, FileTypes.CSV, delimiter
        )


def export_voucher_codes(
    export_file: "ExportFile",
    file_type: str,
    voucher_id: int | None = None,
    ids: list[int] | None = None,
    delimiter: str = ",",
):
    file_name = get_filename("voucher_code", file_type)

    qs = VoucherCode.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME).all()
    if voucher_id:
        qs = VoucherCode.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(voucher_id=voucher_id)
    if ids:
        qs = VoucherCode.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(id__in=ids)

    export_fields = ["code"]
    temporary_file = create_file_with_headers(export_fields, delimiter, file_type)

    export_voucher_codes_in_batches(
        qs,
        export_fields,
        delimiter,
        temporary_file,
        file_type,
    )

    save_csv_file_in_export_file(export_file, temporary_file, file_name)
    temporary_file.close()
    send_export_download_link_notification(export_file, "voucher codes")


def get_filename(model_name: str, file_type: str) -> str:
    hash = uuid.uuid4()
    return "{}_data_{}_{}.{}".format(
        model_name, timezone.now().strftime("%d_%m_%Y_%H_%M_%S"), hash, file_type
    )


def get_queryset(model, filter, scope: dict[str, str | dict]) -> "QuerySet":
    queryset = model.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME).all()
    if "ids" in scope:
        queryset = model.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(pk__in=scope["ids"])
    elif "filter" in scope:
        queryset = filter(data=parse_input(scope["filter"]), queryset=queryset).qs

    queryset = queryset.order_by("pk")

    return queryset


def parse_input(data: Any) -> dict[str, str | dict]:
    """Parse input into correct data types.

    Scope coming from Celery will be passed as strings.
    """
    if "attributes" in data:
        serialized_attributes = []

        for attr in data.get("attributes") or []:
            if "date_time" in attr:
                if gte := attr["date_time"].get("gte"):
                    attr["date_time"]["gte"] = datetime.datetime.fromisoformat(gte)
                if lte := attr["date_time"].get("lte"):
                    attr["date_time"]["lte"] = datetime.datetime.fromisoformat(lte)

            if "date" in attr:
                if gte := attr["date"].get("gte"):
                    attr["date"]["gte"] = datetime.date.fromisoformat(gte)
                if lte := attr["date"].get("lte"):
                    attr["date"]["lte"] = datetime.date.fromisoformat(lte)

            serialized_attributes.append(attr)

        if serialized_attributes:
            data["attributes"] = serialized_attributes

    return data


def format_as_price_list(excel_path: str, export_info: dict[str, list]) -> None:
    """Post-process Excel export into price list format.

    - Renames columns to clean names
    - Removes technical columns (IDs, internal fields)
    - Reorders columns to standard layout
    - Formats RRP with currency symbol

    Args:
        excel_path: Path to the Excel file
        export_info: Export info containing channel IDs and attributes

    """
    from io import BytesIO

    from openpyxl.drawing.image import Image as XLImage

    from ...channel.models import Channel

    wb = load_workbook(excel_path)
    ws = wb.active

    # Extract images BEFORE any transformations
    # Store them by row number so we can reinsert them later
    images_by_row = {}
    if hasattr(ws, "_images") and ws._images:
        logger.info("Extracting %s images before reformatting", len(ws._images))
        for img in ws._images:
            try:
                anchor = img.anchor
                if hasattr(anchor, "_from"):
                    row_idx = anchor._from.row + 1  # Convert to 1-indexed
                    # Store image data and dimensions
                    img_bytes = img._data()
                    images_by_row[row_idx] = {
                        "data": img_bytes,
                        "width": img.width,
                        "height": img.height,
                    }
                    logger.debug("Extracted image from row %s", row_idx)
            except Exception as e:
                logger.warning("Failed to extract image: %s", e)
        logger.info("Extracted %s images from sheet", len(images_by_row))

    # Get current headers
    headers = [cell.value for cell in ws[1]]

    # Column renaming map
    RENAME_MAP = {
        "product media": "Image",
        "name": "Description",
        "category": "Category",
        "variants__size_quantity": "Sizes",
        "variants__total_quantity": "Qty",
    }

    # Add attribute renames (look for patterns)
    for _i, header in enumerate(headers):
        if header:
            header_str = str(header)
            header_lower = header_str.lower()
            # Rename attribute columns (case-insensitive, handle both hyphens and spaces)
            if (
                "product code" in header_lower or "product-code" in header_lower
            ) and "(product attribute)" in header_lower:
                RENAME_MAP[header] = "Product Code"
            elif "brand" in header_lower and "(product attribute)" in header_lower:
                RENAME_MAP[header] = "Brand"
            elif "rrp" in header_lower and "(product attribute)" in header_lower:
                RENAME_MAP[header] = "RRP"

    # Get channel info for price column renaming
    channel_ids = export_info.get("channels")
    channel_slug = None
    currency_code = None

    if channel_ids:
        converted_channel_ids = convert_ids_to_proper_type(channel_ids)
        channels = Channel.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(pk__in=converted_channel_ids)
        if channels.exists():
            channel = channels.first()
            if channel:
                channel_slug = channel.slug
                currency_code = channel.currency_code

    # Rename price columns (remove channel prefix)
    if channel_slug:
        for _i, header in enumerate(headers):
            if header:
                header_str = str(header)
                # Rename "{channel} (channel price amount)" to "Price"
                if f"{channel_slug} (channel price amount)" == header_str:
                    RENAME_MAP[header] = "Price"

    # Columns to remove (by substring match)
    REMOVE_PATTERNS = [
        "id",  # Product/variant IDs
        "product type",
        "(channel variant currency code)",
        "(channel variant cost price)",
        "(channel published)",
        "(channel publication date)",  # Added
        "(channel published at)",  # Added
        "(channel searchable)",
        "(channel available for purchase)",
        "(channel product currency code)",
        "(channel variant preorder quantity threshold)",
    ]

    # Standard column order (columns that exist will be reordered to this)
    COLUMN_ORDER = [
        "Image",
        "Product Code",
        "Description",
        "Category",
        "Sizes",
        "Brand",
        "Qty",
        "RRP",
        "Price",
    ]

    # Step 1: Rename columns
    for cell in ws[1]:
        if cell.value in RENAME_MAP:
            cell.value = RENAME_MAP[cell.value]

    # Step 2: Remove unwanted columns
    headers_after_rename = [cell.value for cell in ws[1]]
    columns_to_delete = []

    for idx, header in enumerate(headers_after_rename, start=1):
        if header:
            header_str = str(header)
            # Check if header matches any remove pattern
            if any(pattern in header_str for pattern in REMOVE_PATTERNS):
                columns_to_delete.append(idx)

    # Delete columns in reverse order to maintain indices
    for col_idx in sorted(columns_to_delete, reverse=True):
        ws.delete_cols(col_idx, 1)

    # Step 3: Reorder columns to standard layout
    final_headers = [cell.value for cell in ws[1]]
    current_order = {header: idx for idx, header in enumerate(final_headers, start=1)}

    # Build new column order based on what exists
    new_order = []
    for col_name in COLUMN_ORDER:
        if col_name in current_order:
            new_order.append((col_name, current_order[col_name]))

    # Add any remaining columns not in COLUMN_ORDER (at the end)
    for header, idx in current_order.items():
        if header not in COLUMN_ORDER:
            new_order.append((header, idx))

    # Reorder by moving columns
    # This is complex in openpyxl, so we'll create a new sheet with correct order

    new_ws = wb.create_sheet("PriceList")

    # Copy headers in new order
    for new_col, (col_name, old_col) in enumerate(new_order, start=1):
        new_ws.cell(row=1, column=new_col).value = col_name

        # Copy all data for this column
        for row_num in range(2, ws.max_row + 1):
            old_cell = ws.cell(row=row_num, column=old_col)
            new_cell = new_ws.cell(row=row_num, column=new_col)
            new_cell.value = old_cell.value
            # Copy number format if it exists
            if old_cell.number_format:
                new_cell.number_format = old_cell.number_format

    # Copy row dimensions (heights) from old sheet to new sheet
    for row_num, row_dim in ws.row_dimensions.items():
        if row_dim.height:
            new_ws.row_dimensions[row_num].height = row_dim.height

    # Copy column dimensions (widths) - need to map old columns to new columns
    for new_col, (_col_name, old_col) in enumerate(new_order, start=1):
        old_col_letter = get_column_letter(old_col)
        new_col_letter = get_column_letter(new_col)
        if old_col_letter in ws.column_dimensions:
            old_col_dim = ws.column_dimensions[old_col_letter]
            if old_col_dim.width:
                new_ws.column_dimensions[new_col_letter].width = old_col_dim.width

    # Delete old sheet and rename new one
    wb.remove(ws)
    new_ws.title = "Sheet"

    # Now reinsert the images into the correct column (Image column)
    # Find the "Image" column in the new sheet
    new_headers = [cell.value for cell in new_ws[1]]
    image_col_idx = None
    for idx, header in enumerate(new_headers, start=1):
        if header == "Image":
            image_col_idx = idx
            break

    if image_col_idx and images_by_row:
        logger.info(
            "Reinserting %s images into column %s (Image)",
            len(images_by_row),
            get_column_letter(image_col_idx),
        )
        image_col_letter = get_column_letter(image_col_idx)

        for row_idx, img_data in images_by_row.items():
            try:
                # Create new image from stored data
                img_bytes = BytesIO(img_data["data"])
                new_img = XLImage(img_bytes)
                new_img.width = img_data["width"]
                new_img.height = img_data["height"]

                # Anchor to the Image column at the same row
                new_anchor = f"{image_col_letter}{row_idx}"
                new_img.anchor = new_anchor

                # Add to worksheet
                new_ws.add_image(new_img)
                logger.debug("Inserted image at %s", new_anchor)
            except Exception as e:
                logger.warning("Failed to insert image at row %s: %s", row_idx, e)

        logger.info("Successfully reinserted images")
    elif images_by_row and not image_col_idx:
        logger.warning(
            "Images were extracted but 'Image' column not found in new sheet"
        )

    # Step 4: Format RRP column with currency (if exists)
    headers_final = [cell.value for cell in new_ws[1]]
    logger.info(
        "RRP formatting check: currency_code=%s, headers=%s",
        currency_code,
        headers_final,
    )

    if currency_code and "RRP" in headers_final:
        # Currency formats
        CURRENCY_FORMATS = {
            "GBP": "[$£-809]#,##0.00",
            "USD": "[$$-409]#,##0.00",
            "EUR": "[$€-407]#,##0.00",
            "JPY": "[$¥-411]#,##0",
            "CNY": "[$¥-804]#,##0.00",
            "INR": "[$₹-439]#,##0.00",
            "AUD": "[$A$-C09]#,##0.00",
            "CAD": "[$C$-1009]#,##0.00",
        }

        number_format = CURRENCY_FORMATS.get(currency_code, "#,##0.00")
        logger.info("Applying currency format %s to RRP column", number_format)

        # Find RRP column
        rrp_col = headers_final.index("RRP") + 1
        formatted_count = 0
        for row in range(2, new_ws.max_row + 1):
            cell = new_ws.cell(row=row, column=rrp_col)
            if cell.value is not None:
                # Convert to float if it's a string
                try:
                    if isinstance(cell.value, str):
                        cell.value = float(cell.value)
                    cell.number_format = number_format
                    formatted_count += 1
                except (ValueError, TypeError):
                    # Keep original value if conversion fails
                    logger.warning(
                        "Could not convert RRP value to number: %s", cell.value
                    )

        logger.info("Formatted %s RRP cells with currency", formatted_count)

    wb.save(excel_path)
    logger.info("Applied price list formatting to export")


def format_currency_columns(excel_path: str, export_info: dict[str, list]) -> None:
    """Format price columns in Excel with currency symbols based on channel currency.

    Args:
        excel_path: Path to the Excel file
        export_info: Export info containing channel IDs

    """
    from ...channel.models import Channel

    channel_ids = export_info.get("channels")
    if not channel_ids:
        return

    # Convert channel IDs to proper type
    converted_channel_ids = convert_ids_to_proper_type(channel_ids)

    # Get channel currencies
    channels = Channel.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME).filter(
        pk__in=converted_channel_ids
    )

    # Currency code to Excel number format mapping
    CURRENCY_FORMATS = {
        "GBP": "[$£-809]#,##0.00",
        "USD": "[$$-409]#,##0.00",
        "EUR": "[$€-407]#,##0.00",
        "JPY": "[$¥-411]#,##0",
        "CNY": "[$¥-804]#,##0.00",
        "INR": "[$₹-439]#,##0.00",
        "AUD": "[$A$-C09]#,##0.00",
        "CAD": "[$C$-1009]#,##0.00",
        "CHF": "[$CHF-807]#,##0.00",
        "SEK": "[$kr-41D]#,##0.00",
        "NOK": "[$kr-414]#,##0.00",
        "DKK": "[$kr-406]#,##0.00",
        "PLN": "[$zł-415]#,##0.00",
    }

    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Track total columns formatted
    total_columns_formatted = 0

    # Find price-related columns for each channel
    for channel in channels:
        currency_code = channel.currency_code
        channel_slug = channel.slug
        number_format = CURRENCY_FORMATS.get(currency_code, "#,##0.00")

        # Find columns that contain price amounts for this channel
        price_columns = []
        for idx, header in enumerate(headers):
            if header and channel_slug in str(header):
                # Match patterns like "default (channel price amount)"
                if "price amount" in str(header).lower():
                    price_columns.append(idx + 1)  # 1-indexed
                # Also format cost price
                elif "cost price" in str(header).lower():
                    price_columns.append(idx + 1)

        # Apply formatting to these columns
        for col_idx in price_columns:
            for row in range(2, ws.max_row + 1):  # Skip header
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = number_format

        total_columns_formatted += len(price_columns)

    # Save workbook
    wb.save(excel_path)
    logger.info(
        "Applied currency formatting to %s price columns", total_columns_formatted
    )


def create_file_with_headers(file_headers: list[str], delimiter: str, file_type: str):
    table = etl.wrap([file_headers])

    if file_type == FileTypes.CSV:
        temp_file = NamedTemporaryFile("ab+", suffix=".csv")
        etl.tocsv(table, temp_file.name, delimiter=delimiter)
    else:
        temp_file = NamedTemporaryFile("ab+", suffix=".xlsx")
        etl.io.xlsx.toxlsx(table, temp_file.name)

    return temp_file


def export_products_in_batches(
    queryset: "QuerySet",
    export_info: dict[str, list],
    export_fields: set[str],
    headers: list[str],
    delimiter: str,
    temporary_file: Any,
    file_type: str,
):
    warehouses = export_info.get("warehouses")
    attributes = export_info.get("attributes")
    channels = export_info.get("channels")

    for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
        product_batch = (
            Product.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME)
            .filter(pk__in=batch_pks)
            .prefetch_related(
                "attributevalues",
                "variants",
                "collections",
                "media",
                "product_type",
                "category",
            )
        )
        export_data = get_products_data(
            product_batch, export_fields, attributes, warehouses, channels
        )

        append_to_file(
            cast(list[dict[str, str | bool | float | int | None]], export_data),
            headers,
            temporary_file,
            file_type,
            delimiter,
        )


def export_products_compressed(
    queryset: "QuerySet",
    export_info: dict[str, list],
    headers: list[str],
    delimiter: str,
    temporary_file: Any,
    file_type: str,
):
    """Export products with compressed variants (one row per product)."""
    from collections import ChainMap

    from . import ProductExportFields

    warehouses = export_info.get("warehouses")
    attributes = export_info.get("attributes")
    channels = export_info.get("channels")
    requested_fields = export_info.get("fields", [])

    # Build export_fields set from requested fields
    fields_mapping = dict(
        ChainMap(*reversed(ProductExportFields.HEADERS_TO_FIELDS_MAPPING.values()))
    )
    export_fields_set = set()
    if requested_fields:
        for field in requested_fields:
            if field in fields_mapping:
                export_fields_set.add(fields_mapping[field])

    for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
        product_batch = (
            Product.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME)
            .filter(pk__in=batch_pks)
            .prefetch_related(
                "attributevalues",
                "variants",
                "collections",
                "media",
                "product_type",
                "category",
            )
        )

        # Use compressed data function
        export_data = get_products_data_compressed(
            product_batch,
            export_fields_set,  # Pass proper export fields
            attributes,
            warehouses,
            channels,
            requested_fields,
        )

        append_to_file(export_data, headers, temporary_file, file_type, delimiter)


def export_gift_cards_in_batches(
    queryset: "QuerySet",
    export_fields: list[str],
    delimiter: str,
    temporary_file: Any,
    file_type: str,
):
    for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
        gift_card_batch = GiftCard.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(pk__in=batch_pks)

        export_data = list(gift_card_batch.values(*export_fields))

        append_to_file(export_data, export_fields, temporary_file, file_type, delimiter)


def export_voucher_codes_in_batches(
    queryset: "QuerySet",
    export_fields: list[str],
    delimiter: str,
    temporary_file: Any,
    file_type: str,
):
    for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
        voucher_codes_batch = VoucherCode.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(pk__in=batch_pks)

        export_data = list(voucher_codes_batch.values(*export_fields))

        append_to_file(export_data, export_fields, temporary_file, file_type, delimiter)


def append_to_file(
    export_data: list[dict[str, str | bool | float | int | None]],
    headers: list[str],
    temporary_file: Any,
    file_type: str,
    delimiter: str,
):
    table = etl.fromdicts(export_data, header=headers, missing="")

    if file_type == FileTypes.CSV:
        etl.io.csv.appendcsv(table, temporary_file.name, delimiter=delimiter)
    else:
        etl.io.xlsx.appendxlsx(table, temporary_file.name)


@allow_writer()
def save_csv_file_in_export_file(
    export_file: "ExportFile", temporary_file: IO[bytes], file_name: str
):
    export_file.content_file.save(file_name, temporary_file)


PO_SUMMARY_HEADERS = [
    "PO Name",
    "Status",
    "Supplier",
    "Destination",
    "Currency",
    "Created",
    "Total Ordered Qty",
    "Total Received Qty",
    "Total Value",
]

PO_LINE_HEADERS = [
    "PO Name",
    "Product Name",
    "Product Code",
    "SKU",
    "Ordered",
    "Shipped",
    "Received",
    "Difference",
    "Unit Price",
    "Total Price",
    "Currency",
    "Country of Origin",
    "Status",
]

PO_EXPORT_HEADERS = PO_SUMMARY_HEADERS + [h for h in PO_LINE_HEADERS if h != "PO Name"]


def _compute_unit_price(poi) -> Decimal | int:
    from decimal import Decimal

    if poi.quantity_ordered == 0 or poi.total_price_amount is None:
        return 0
    base_unit_price = poi.total_price_amount / poi.quantity_ordered

    payable_adj = sum(
        a.quantity_change
        for a in poi.adjustments.all()
        if a.affects_payable and a.processed_at is not None
    )
    adjusted_cost = poi.total_price_amount + (Decimal(payable_adj) * base_unit_price)

    all_adj = sum(
        a.quantity_change for a in poi.adjustments.all() if a.processed_at is not None
    )
    adjusted_quantity = poi.quantity_ordered + all_adj

    if adjusted_quantity > 0:
        return adjusted_cost / adjusted_quantity
    return base_unit_price


def _get_variant_size(variant, size_values_map: dict) -> str:
    size = size_values_map.get(variant.pk)
    if size:
        return size
    return variant.sku or f"V{variant.pk}"


def _build_size_qty_string(size_qty_pairs: list[tuple[str, int]]) -> str:
    try:
        sorted_pairs = sorted(size_qty_pairs, key=lambda x: float(x[0]))
    except (ValueError, TypeError):
        sorted_pairs = sorted(size_qty_pairs, key=lambda x: str(x[0]))
    return ", ".join(f"{size}[{qty}]" for size, qty in sorted_pairs)


def _build_po_rows(
    po, product_code_slug: str, size_values_map: dict
) -> tuple[dict, list[dict]]:
    from collections import defaultdict
    from decimal import Decimal

    supplier = po.source_warehouse.name if po.source_warehouse else ""
    destination = po.destination_warehouse.name if po.destination_warehouse else ""

    total_ordered = 0
    total_received = 0
    total_value = Decimal(0)

    product_groups: dict[int, dict] = defaultdict(
        lambda: {
            "product_name": "",
            "product_code": "",
            "sku": "",
            "ordered": [],
            "shipped": [],
            "received": [],
            "unit_prices": [],
            "total_prices": [],
            "currencies": [],
            "countries": [],
            "statuses": [],
        }
    )

    for poi in po.items.all():
        variant = poi.product_variant
        product = variant.product
        product_id = product.pk
        size = _get_variant_size(variant, size_values_map)

        group = product_groups[product_id]
        if not group["product_name"]:
            group["product_name"] = product.name
            group["sku"] = variant.sku or ""
            for av in product.attributevalues.all():
                if av.value.attribute.slug == product_code_slug:
                    group["product_code"] = av.value.name
                    break

        group["ordered"].append((size, poi.quantity_ordered))
        total_ordered += poi.quantity_ordered

        is_shipped = poi.shipment_id and poi.shipment and poi.shipment.departed_at
        group["shipped"].append((size, poi.quantity_ordered) if is_shipped else None)

        receipt_lines = list(poi.receipt_lines.all())
        if receipt_lines:
            received_qty = sum(rl.quantity_received for rl in receipt_lines)
            group["received"].append((size, received_qty))
            total_received += received_qty
        else:
            group["received"].append(None)

        if poi.total_price_amount is not None:
            total_value += poi.total_price_amount

        unit_price = _compute_unit_price(poi)
        if unit_price and str(round(unit_price, 2)) not in group["unit_prices"]:
            group["unit_prices"].append(str(round(unit_price, 2)))
        if poi.total_price_amount is not None:
            group["total_prices"].append(poi.total_price_amount)
        if poi.currency and poi.currency not in group["currencies"]:
            group["currencies"].append(poi.currency)
        coo = str(poi.country_of_origin) if poi.country_of_origin else ""
        if coo and coo not in group["countries"]:
            group["countries"].append(coo)
        if poi.status and poi.status not in group["statuses"]:
            group["statuses"].append(poi.status)

    summary = {
        "PO Name": po.name or str(po.pk),
        "Status": po.status,
        "Supplier": supplier,
        "Destination": destination,
        "Currency": po.currency or "",
        "Created": po.created_at.strftime("%Y-%m-%d") if po.created_at else "",
        "Total Ordered Qty": total_ordered,
        "Total Received Qty": total_received,
        "Total Value": str(total_value),
    }

    line_rows: list[dict] = []
    for _product_id, group in product_groups.items():
        ordered_pairs = group["ordered"]
        shipped_actual = [p for p in group["shipped"] if p is not None]
        received_actual = [p for p in group["received"] if p is not None]

        diff_pairs = []
        if received_actual:
            for i, (size, ordered_qty) in enumerate(ordered_pairs):
                recv = group["received"][i]
                if recv is not None:
                    d = recv[1] - ordered_qty
                    if d != 0:
                        diff_pairs.append((size, d))

        line_total = sum(group["total_prices"])

        line_rows.append(
            {
                "PO Name": po.name or str(po.pk),
                "Product Name": group["product_name"],
                "Product Code": group["product_code"],
                "SKU": group["sku"],
                "Ordered": _build_size_qty_string(ordered_pairs),
                "Shipped": _build_size_qty_string(shipped_actual)
                if shipped_actual
                else "",
                "Received": _build_size_qty_string(received_actual)
                if received_actual
                else "",
                "Difference": _build_size_qty_string(diff_pairs) if diff_pairs else "",
                "Unit Price": ", ".join(group["unit_prices"]),
                "Total Price": str(line_total) if line_total else "",
                "Currency": ", ".join(group["currencies"]),
                "Country of Origin": ", ".join(group["countries"]),
                "Status": ", ".join(group["statuses"]),
            }
        )

    return summary, line_rows


def _po_batch_queryset(batch_pks):
    return (
        PurchaseOrder.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME)
        .filter(pk__in=batch_pks)
        .select_related(
            "source_warehouse",
            "destination_warehouse",
        )
        .prefetch_related(
            "items__product_variant__product__attributevalues__value__attribute",
            "items__shipment",
            "items__receipt_lines",
            "items__adjustments",
        )
        .order_by("pk")
    )


def _get_size_values_for_variants(variant_ids: list) -> dict:
    if not variant_ids:
        return {}

    from ...attribute.models import AssignedVariantAttribute

    size_values_map: dict = {}
    for slug in ["size", "Size", "SIZE"]:
        assigned_attrs = (
            AssignedVariantAttribute.objects.using(
                settings.DATABASE_CONNECTION_REPLICA_NAME
            )
            .filter(
                variant_id__in=variant_ids,
                assignment__attribute__slug=slug,
            )
            .prefetch_related("values")
        )
        for assigned_attr in assigned_attrs:
            values = assigned_attr.values.all()
            if values:
                size_values_map[assigned_attr.variant_id] = (
                    values[0].name or values[0].slug or str(values[0].value)
                )
        if size_values_map:
            break
    return size_values_map


def export_purchase_orders(
    export_file: "ExportFile",
    scope: dict[str, str | dict],
    file_type: str,
    delimiter: str = ",",
):
    from ...graphql.inventory.filters import PurchaseOrderFilter
    from ...site.models import SiteSettings

    file_name = get_filename("purchase_order", file_type)
    queryset = get_queryset(PurchaseOrder, PurchaseOrderFilter, scope)

    site_settings = SiteSettings.objects.first()
    product_code_slug = (
        site_settings.invoice_product_code_attribute
        if site_settings
        else "product-code"
    )

    all_variant_ids = list(
        PurchaseOrder.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME)
        .filter(pk__in=queryset.values_list("pk", flat=True))
        .values_list("items__product_variant_id", flat=True)
        .distinct()
    )
    size_values_map = _get_size_values_for_variants(
        [v for v in all_variant_ids if v is not None]
    )

    if file_type == FileTypes.XLSX:
        summary_rows: list[dict] = []
        lines_rows: list[dict] = []
        for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
            for po in _po_batch_queryset(batch_pks):
                summary, line_rows = _build_po_rows(
                    po, product_code_slug, size_values_map
                )
                summary_rows.append(summary)
                lines_rows.extend(line_rows)

        temp_file = NamedTemporaryFile("ab+", suffix=".xlsx")
        _write_po_xlsx(summary_rows, lines_rows, temp_file.name)
        temporary_file = temp_file
    else:
        temporary_file = create_file_with_headers(
            PO_EXPORT_HEADERS, delimiter, file_type
        )
        for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
            export_data: list[dict] = []
            for po in _po_batch_queryset(batch_pks):
                summary, line_rows = _build_po_rows(
                    po, product_code_slug, size_values_map
                )
                if not line_rows:
                    export_data.append(
                        {
                            **summary,
                            **{h: "" for h in PO_LINE_HEADERS if h != "PO Name"},
                        }
                    )
                else:
                    for line_row in line_rows:
                        export_data.append(
                            {
                                **summary,
                                **{k: v for k, v in line_row.items() if k != "PO Name"},
                            }
                        )
            append_to_file(
                export_data, PO_EXPORT_HEADERS, temporary_file, FileTypes.CSV, delimiter
            )

    save_csv_file_in_export_file(export_file, temporary_file, file_name)
    temporary_file.close()
    send_export_download_link_notification(export_file, "purchase orders")


def _write_po_xlsx(summary_rows: list[dict], lines_rows: list[dict], filename: str):
    wb = openpyxl.Workbook()

    ws_po = wb.active
    ws_po.title = "Purchase Orders"
    ws_po.append(PO_SUMMARY_HEADERS)
    for row in summary_rows:
        ws_po.append([row.get(h, "") for h in PO_SUMMARY_HEADERS])

    ws_lines = wb.create_sheet("Lines")
    ws_lines.append(PO_LINE_HEADERS)
    for row in lines_rows:
        ws_lines.append([row.get(h, "") for h in PO_LINE_HEADERS])

    wb.save(filename)
